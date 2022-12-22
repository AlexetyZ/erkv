from typing import Type, TypeVar
from pathlib import Path
from erknm.pxl import Exel_work
import openpyxl
from openpyxl.styles import PatternFill


Workbook = TypeVar('Workbook', bound=openpyxl.Workbook)

class Operation:
    def __init__(self, wb_path: str = 'Новый документ.xlsx', init_sh: bool = True):
        """
        :param wb_path: путь до файла, в котором итерируются ячейки
        :param init_sh: нужно ли инициировать первый лист книги. По умолчанию True. Необходимо для ситуаций,
        когда требуется сберечь вычислительный ресурс при отсутствии необходимости работы с листом книги,
        например открытия книги для сохранения в другом формате, без обращения к листу.
        """
        self.wb_path = wb_path
        if self.wb_path == 'Новый документ.xlsx':
            self.wb = openpyxl.Workbook()
        else:
            self.wb = openpyxl.load_workbook(wb_path)
        if init_sh is True:
            self.sh = self.wb.worksheets[0]


    def get_list_from_sh_column(self, *columns: str, start_from_row: int = 1, reference_column: str = 'A',
                                del_last_empty_rows=False):
        """
        Если нужно получить значения определенного столбца в файле exel.


        :param wb_path: путь до файла, в котором итерируются ячейки
        :param column: столбец, в котором итерируются ячейки
        :param start_from_row: начало итерации для удаления пустых строк
        :param reference_column: показательный столбец, по которому отсчитывается количество строк, то есть,
            не пустое значение ячейки этого столбца гарантирует, что вся строка подлежит отработке. По умолчанию "А"
        :param del_last_empty_rows: bool При значении True отрезает пустые строки снизу
        :return: Возвращает список значений столбца в ячейке exel
        """
        list = []
        if del_last_empty_rows is True:
            Exel_work().delete_last_empty_rows(self.wb_path, start_from_row)

        for n, row in enumerate(self.sh[reference_column]):
            if n + 1 < start_from_row:
                continue
            corteg = []
            for column in columns:
                corteg.append(self.sh[f'{column}{n + 1}'].value)
            list.append(tuple(corteg))
        return list

    def get_column_values(self, column: str) -> list:
        result = []
        list = self.sh[column]
        for l in list:
            result.append(l.value)
        return result

    def return_path_file(self):
        return self.wb_path

    def create_doc_in_this_path(self, name):
        wb_0 = openpyxl.Workbook()
        path = Path(self.wb_path).parent
        pathname = f'{path}/{name}.xlsx'
        self.save_document(workbook=wb_0, path=pathname)
        return pathname

    def get_cell_value(self, row: int, column: int):
        return self.sh.cell(row=row, column=column).value

    def change_value_in_cell(self, row: int, column: int, value, number_format: str = 'no', saving: bool = True):
        if number_format != 'no':
            self.sh.cell(row=row, column=column).number_format = number_format
        self.sh.cell(row=row, column=column, value=value)
        if saving is True:
            self.save_document()


    def mark_cell(self, row: int, column: int, color: str = "ffff00", saving: bool = True):
        self.sh.cell(row=row, column=column).fill = PatternFill('solid', fgColor=color)
        if saving is True:
            self.save_document()

    def save_document(self, workbook: Type[Workbook] or bool = False, path: str or bool = False):
        if path is False:
            path = self.wb_path
        else:
            path += '.xlsx'
        if workbook is False:
            workbook = self.wb
        workbook.save(path)

def main():
    operation_1 = Operation("C:\\Users\zaitsev_ad\Desktop\Список_планов_КНМ (2).xlsx")
    operation_1.create_doc_in_this_path('yjdsq')


if __name__ == '__main__':
    main()
